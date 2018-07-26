# encoding = utf-8
from openpyxl import load_workbook


class ParseExcel(object):

    def __init__(self, excelPath, sheetName):
        self.lwb = load_workbook(excelPath)
        #  self.sheet = self.lwb.get_sheet_by_name(sheetName)
        self.sheet = self.lwb[sheetName]
        self.maxRowNum = self.sheet.max_row

    def getDatasFromSheet(self):
        for line in self.sheet.rows:
            for cell in line:
                print(cell.value)


if __name__ == '__main__':
    excelPath = u'F:\\seleniumWithPython\\TestData\\测试数据.xlsx'
    sheetName = u'搜索数据表'
    pe = ParseExcel(excelPath, sheetName)
    pe.getDatasFromSheet()
